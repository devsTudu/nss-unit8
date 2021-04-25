from openpyxl import load_workbook
from models import Members
  

workbook = load_workbook(filename="infomembers.xlsx")

sheet = workbook.active

for i in range(40):
  print(sheet.cell(4+i,3).value)
