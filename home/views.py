from django.shortcuts import render, HttpResponse,redirect

from openpyxl import load_workbook

from .forms import CommentForm
from .models import Post, Members

def homepage(request):
  posts = Post.objects.all()
  return render(request,'index.html',{'posts':posts})

def post_detail(request, slug):
    post = Post.objects.get(slug=slug)

    if request.method == 'POST':
        form = CommentForm(request.POST)

        if form.is_valid():
            comment = form.save(commit=False)
            comment.post = post
            comment.save()

            return redirect('post_detail', slug=post.slug)
    else:
        form = CommentForm()

    return render(request, 'readblog.html', {'post': post, 'form': form})

def updateMembers(request):
  Members.objects.all().delete()
  workbook = load_workbook(filename="membersinfo.xlsx")
  sheet = workbook.active
  for i in range(40):
    n = sheet.cell(4+i,3).value
    e = str(sheet.cell(4+i,4).value)
    r = sheet.cell(4+i,2).value
    c = sheet.cell(4+i,9).value
    img = sheet.cell(4+i,10).value
    if not img or len(str(img))<5 :
      img = ''
    else:
      img = str(img).replace("/open?","/uc?")
    obj = Members.objects.create(Name=n,Email=e,Roll_Number=r,image_url=img, city=c)
    obj.save()
  return redirect('members/')

def seeMembers(request):
  members = Members.objects.all()
  return render(request,'members.html',{'members':members})
